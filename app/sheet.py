# SPDX-License-Identifier: MPL-2.0
"""Reading a spreadsheet as rows, and writing verdicts back out.

openpyxl is already here for indexing .xlsx, so this borrows it rather than
adding a dependency. Two things it does differently from the indexer, both
because the purpose differs: it reads *cells by column name* rather than
flattening a sheet to text, and it streams in read-only mode, because a
16,000-row workbook loaded whole is a few hundred megabytes of Python objects
for data that is mostly short strings.
"""

from __future__ import annotations

import csv
import io
import logging
import os

log = logging.getLogger("rag.sheet")

MAX_ROWS = 200_000


def read_rows(path: str, sheet: str | None = None, limit: int = 0) -> tuple[list[dict], list[str]]:
    """(rows as dicts keyed by header, the header names in order)."""
    extension = os.path.splitext(path)[1].lower()
    if extension in (".csv", ".tsv"):
        return _read_delimited(path, limit)
    if extension not in (".xlsx", ".xlsm"):
        raise ValueError(f"cannot read {extension or 'a file with no extension'} - "
                         f"use .xlsx, .xlsm, .csv or .tsv")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("openpyxl is not installed") from exc

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        try:
            page = book[sheet] if sheet else book[book.sheetnames[0]]
        except KeyError:
            raise ValueError(
                f"no sheet named {sheet!r} - this workbook has: "
                f"{', '.join(book.sheetnames)}"
            )
        stream = page.iter_rows(values_only=True)
        try:
            header_row = next(stream)
        except StopIteration:
            return [], []
        headers = [
            (str(value).strip() if value is not None else f"column{index + 1}")
            for index, value in enumerate(header_row)
        ]
        rows: list[dict] = []
        for values in stream:
            if limit and len(rows) >= limit:
                break
            if len(rows) >= MAX_ROWS:
                log.warning("%s: stopping at %d rows", path, MAX_ROWS)
                break
            # A wholly empty row is spreadsheet padding, not a record.
            if not any(value is not None and str(value).strip() for value in values):
                continue
            rows.append({
                headers[index]: value
                for index, value in enumerate(values)
                if index < len(headers)
            })
        return rows, headers
    finally:
        book.close()


def _read_delimited(path: str, limit: int) -> tuple[list[dict], list[str]]:
    delimiter = "\t" if path.lower().endswith(".tsv") else ","
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        headers = list(reader.fieldnames or [])
        rows = []
        for row in reader:
            if limit and len(rows) >= limit:
                break
            if any((value or "").strip() for value in row.values()):
                rows.append(row)
    return rows, headers


def to_csv(verdicts: list[dict], rows: list[dict], carry: list[str]) -> str:
    """Verdicts as CSV, carrying through the columns that identify a row.

    Written as CSV rather than back into the workbook on purpose: editing
    someone's source spreadsheet in place is not a thing a tool should do
    unasked, and a separate file can be compared, re-run and thrown away.
    """
    buffer = io.StringIO()
    columns = ["row"] + list(carry) + ["label", "criterion", "confidence", "error"]
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for index, verdict in enumerate(verdicts):
        record = {key: rows[index].get(key) for key in carry} if index < len(rows) else {}
        record.update(verdict)
        writer.writerow(record)
    return buffer.getvalue()
